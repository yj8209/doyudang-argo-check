import streamlit as st
import pandas as pd
from PIL import Image
import io
import os
from datetime import datetime
from openpyxl.utils import get_column_letter

# --- 페이지 설정 및 로고 ---
st.set_page_config(page_title="두유당 ARGO 정산 검증 대시보드", layout="wide")

# 로고 이미지 로드
logo_path = "static/KakaoTalk_20260405_223421313.png" 
try:
    image = Image.open(logo_path)
    st.image(image, width=200)
except FileNotFoundError:
    st.warning(f"로고 이미지를 찾을 수 없습니다. 경로를 확인해 주세요: {logo_path}")

st.title("두유당 ARGO 월별 정산 정밀 검증 시스템")
st.markdown("""
<style>
    .reportview-container .main .block-container{
        max-width: 1200px;
        padding-top: 2rem;
        padding-right: 2rem;
        padding-left: 2rem;
        padding-bottom: 2rem;
    }
    h1 { color: #1E3A8A; }
    h2 { color: #2563EB; }
    .stFileUploader { padding-bottom: 2rem; border-bottom: 2px solid #E5E7EB; }
</style>
""", unsafe_allow_html=True)

# --- [영구 저장소 설정] CSV 파일 경로 (배상금 관리용) ---
DB_FILE = "compensation_data.csv"

def load_comp_data():
    if os.path.exists(DB_FILE):
        return pd.read_csv(DB_FILE, dtype={'주문번호': str})
    else:
        return pd.DataFrame(columns=[
            "주문번호", "접수일", "처리일", "스토어", "수량", 
            "판매가", "박스수", "합포장", "상품배상금", "택배배상비", "총 배상청구액"
        ])

def save_comp_data(df):
    df.to_csv(DB_FILE, index=False, encoding='utf-8-sig')

if 'compensation_df' not in st.session_state:
    st.session_state.compensation_df = load_comp_data()

# --- 메인 엑셀 파일 업로드 영역 ---
st.subheader("📁 아르고 정산 엑셀 파일 업로드")
uploaded_excel = st.file_uploader("당월 아르고 정산 원본 엑셀 파일(.xlsx)을 이곳에 첨부해 주세요.", type=['xlsx', 'xls'])

# --- 데이터 전처리 헬퍼 함수 ---
def load_excel_sheet(excel_file, sheet_name, skip_rows):
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, skiprows=skip_rows, dtype=str)
        if '고객사명' not in df.columns and len(df.columns) > 0:
            df.columns = df.iloc[0]
            df = df[1:].reset_index(drop=True)
        return df
    except ValueError:
        st.error(f"엑셀 파일 내에 '{sheet_name}' 시트가 존재하지 않습니다.")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"데이터를 읽는 중 오류가 발생했습니다: {e}")
        return pd.DataFrame()

# --- 탭 구성 ---
tab1, tab2, tab3, tab4 = st.tabs(["📊 시스템 개요", "📦 입고비 검증", "🚚 출고 배송비 검증", "💰 배상금 정산 관리"])

# ==========================================
# TAB 1: 시스템 개요
# ==========================================
with tab1:
    st.header("시스템 개요")
    st.write("""
    독립 기업인 두유당의 자체적인 기준에 맞추어, 아르고(ARGO) 풀필먼트에서 매월 청구하는 정산 엑셀 데이터의 오류를 정밀하게 검증하는 시스템입니다.
    
    * **이용 방법:** 상단에 아르고에서 전달받은 원본 엑셀 파일 하나만 업로드하면, 아래 각 탭에서 해당 월의 내역을 자동으로 분석합니다.
    * **배상금 관리:** 파손/오염 건에 대한 내역을 영구적으로 기록하고 관리할 수 있습니다.
    * **출력 기능:** 검증 완료 후 식별된 오류 내역은 엑셀 파일로 다운로드하여 증빙 자료로 활용할 수 있습니다.
    """)

# ==========================================
# TAB 2: 입고비 검증
# ==========================================
with tab2:
    st.header("입고비 정밀 검증")
    
    if uploaded_excel is None:
        st.info("👆 상단에 엑셀 파일을 먼저 업로드해 주세요.")
    else:
        sku_list = ['하루두유 BLACK', '하루두유 BLACK SWEET', '기타 (직접 입력)']
        selected_sku = st.selectbox("검증할 SKU를 선택하세요:", sku_list)
        if selected_sku == '기타 (직접 입력)':
            selected_sku = st.text_input("SKU 이름을 정확히 입력해 주세요:")
            
        actual_inbound_qty = st.number_input("해당 월의 실제 입고 수량을 기입해 주세요:", min_value=0, step=1)
        
        if st.button("입고비 검증 실행"):
            df_in = load_excel_sheet(uploaded_excel, sheet_name='입고비', skip_rows=6)
            
            if not df_in.empty:
                col_idx_in_qty, col_idx_in_amt = 10, 9
                for i, col_name in enumerate(df_in.columns):
                    if '입고 검수비(기본)' in str(col_name):
                        sub_val = str(df_in.iloc[0, i]).strip()
                        if sub_val == '개수': col_idx_in_qty = i
                        elif sub_val == '금액': col_idx_in_amt = i
                
                df_filtered = df_in[df_in['SKU 이름'] == selected_sku]
                
                if df_filtered.empty:
                    st.warning(f"업로드된 데이터에 '{selected_sku}' 입고 내역이 존재하지 않습니다.")
                else:
                    df_filtered['입고유형'] = df_filtered['입고유형'].fillna('')
                    
                    total_billed_qty = 0
                    calculated_total_amt = 0
                    billed_total_amt = 0
                    
                    for index, row in df_filtered.iterrows():
                        if str(row['고객사명']).strip() == '' or str(row.iloc[col_idx_in_qty]).strip() == '개수':
                            continue 
                            
                        try:
                            row_qty_str = str(row.iloc[col_idx_in_qty]).strip()
                            row_billed_amt_str = str(row.iloc[col_idx_in_amt]).strip()
                            
                            row_qty = float(row_qty_str) if row_qty_str.lower() != 'nan' and row_qty_str != '' else 0
                            row_billed_amt = float(row_billed_amt_str) if row_billed_amt_str.lower() != 'nan' and row_billed_amt_str != '' else 0
                            
                            if pd.notna(row_qty):
                                total_billed_qty += row_qty
                                if '당일' in str(row['입고유형']):
                                    calculated_total_amt += (row_qty * 200)
                                else:
                                    calculated_total_amt += (row_qty * 100)
                                    
                            if pd.notna(row_billed_amt):
                                billed_total_amt += row_billed_amt
                        except:
                            continue
                    
                    st.subheader(f"[{selected_sku}] 입고 검증 결과")
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("입력된 실제 입고 수량", f"{actual_inbound_qty:,} 개")
                        st.metric("아르고 청구 입고 수량", f"{int(total_billed_qty):,} 개")
                    with col2:
                        st.metric("아르고 청구 금액", f"{int(billed_total_amt):,} 원")
                        st.metric("자체 산정 금액", f"{int(calculated_total_amt):,} 원")
                    
                    if actual_inbound_qty == total_billed_qty and billed_total_amt == calculated_total_amt:
                        st.success("✅ 수량 및 청구 금액이 모두 정확히 일치합니다.")
                    else:
                        st.error("❌ 수량 또는 금액에 불일치가 발생했습니다. 확인이 필요합니다.")

# ==========================================
# TAB 3: 출고 배송비 검증
# ==========================================
with tab3:
    st.header("출고 배송비 정밀 검증")
    
    if uploaded_excel is None:
        st.info("👆 상단에 엑셀 파일을 먼저 업로드해 주세요.")
    else:
        if st.button("출고 배송비 검증 실행"):
            df_out = load_excel_sheet(uploaded_excel, sheet_name='출고 배송비', skip_rows=4)
            
            if not df_out.empty:
                errors_list = []
                warnings_list = []
                
                col_idx_total, col_idx_island, col_idx_same, col_idx_diff = 14, 19, 15, 16
                for i, val in enumerate(df_out.iloc[0]):
                    val_str = str(val).strip()
                    if val_str == '총 금액': col_idx_total = i
                    elif val_str == '도서 산간 추가 택배비': col_idx_island = i
                    elif val_str == '합포장(동종)': col_idx_same = i
                    elif val_str == '합포장(이종)': col_idx_diff = i
                
                for index, row in df_out.iterrows():
                    if index == 0: continue
                    
                    try:
                        sku_count_str = str(row['SKU 개수']).strip()
                        if sku_count_str.lower() == 'nan' or sku_count_str == '': continue
                        sku_count = int(float(sku_count_str))
                    except:
                        continue 
                    
                    order_number = str(row['주문번호']).replace('.0', '').strip()
                    if order_number.lower() == 'nan': order_number = ""
                    
                    store_name = str(row['스토어명']).strip()
                    actual_grade = str(row['등급']).strip()
                    
                    billed_total_str = str(row.iloc[col_idx_total]).replace(',', '').strip()
                    billed_total = float(billed_total_str) if billed_total_str.lower() != 'nan' and billed_total_str != '' else 0
                    
                    island_val = str(row.iloc[col_idx_island]).strip()
                    island_cost = 3000 if island_val.lower() != 'nan' and island_val != '' else 0
                    
                    same_val = str(row.iloc[col_idx_same]).strip()
                    has_same = same_val.lower() != 'nan' and same_val != ''
                    
                    diff_val = str(row.iloc[col_idx_diff]).strip()
                    has_diff = diff_val.lower() != 'nan' and diff_val != ''
                    
                    is_naver = '네이버스마트스토어' in store_name
                    
                    expected_grade = ""
                    base_shipping = 0
                    box_cost = 0
                    packing_cost = 0
                    error_reasons = []
                    
                    if sku_count >= 8:
                        warnings_list.append({
                            '엑셀 행 번호': index + 6, 
                            '주문번호': order_number,
                            '스토어명': store_name,
                            'SKU 개수': sku_count,
                            '실제 등급': actual_grade,
                            '청구 총금액': billed_total
                        })
                        continue
                        
                    elif sku_count == 1:
                        expected_grade = "극소"
                        base_shipping = 3050 if is_naver else 2750
                        box_cost = 220
                    elif sku_count == 2:
                        expected_grade = "소"
                        base_shipping = 3600 if is_naver else 3300
                        box_cost = 220
                        if has_diff: packing_cost = 100
                        elif has_same: packing_cost = 50
                    elif sku_count in [3, 4]:
                        expected_grade = "중"
                        base_shipping = 4100 if is_naver else 3800
                        box_cost = 450
                        if has_diff: packing_cost = 250
                        elif has_same: packing_cost = 150
                    elif sku_count == 5:
                        expected_grade = "대"
                        base_shipping = 5500 if is_naver else 5000
                        box_cost = 800
                        if has_diff: packing_cost = 250
                        elif has_same: packing_cost = 200
                    elif sku_count in [6, 7]:
                        expected_grade = "특대"
                        base_shipping = 6300 if is_naver else 5800
                        box_cost = 800
                        if has_diff: packing_cost = 400
                        elif has_same: 
                            packing_cost = 250 if sku_count == 6 else 300
                    
                    if actual_grade != expected_grade:
                        error_reasons.append(f"등급 오분류({expected_grade}↔{actual_grade})")
                    
                    expected_total = base_shipping + box_cost + packing_cost + island_cost
                    
                    if billed_total > expected_total:
                        error_reasons.append("금액 초과 청구")
                    
                    if error_reasons:
                        errors_list.append({
                            '엑셀 행': index + 6,
                            '주문번호': order_number,
                            'SKU 개수': sku_count,
                            '오류 사유': " / ".join(error_reasons),
                            '청구 총금액': billed_total,
                            '산정 총금액': expected_total,
                            '초과 청구액': billed_total - expected_total if billed_total > expected_total else 0
                        })
                
                st.subheader("⚠️ 정산 오류 식별 내역 (등급 오분류 및 초과 청구)")
                if errors_list:
                    df_errors = pd.DataFrame(errors_list)
                    
                    total_billed = df_errors['청구 총금액'].sum()
                    total_expected = df_errors['산정 총금액'].sum()
                    total_excess = df_errors['초과 청구액'].sum()

                    total_row = pd.DataFrame({
                        '엑셀 행': [''],
                        '주문번호': ['[ 총 합 계 ]'],
                        'SKU 개수': [''],
                        '오류 사유': [''],
                        '청구 총금액': [total_billed],
                        '산정 총금액': [total_expected],
                        '초과 청구액': [total_excess]
                    })
                    
                    df_errors = pd.concat([df_errors, total_row], ignore_index=True)
                    
                    styled_errors = df_errors.style \
                        .set_table_styles([
                            {'selector': 'th', 'props': [('text-align', 'center')]},
                            {'selector': 'td', 'props': [('text-align', 'center')]}
                        ]) \
                        .set_properties(subset=['주문번호'], **{'text-align': 'right'}) \
                        .set_properties(subset=['초과 청구액'], **{'background-color': '#FFF2CC', 'color': '#D32F2F', 'font-weight': 'bold'}) \
                        .format({'청구 총금액': '{:,.0f}', '산정 총금액': '{:,.0f}', '초과 청구액': '{:,.0f}'})
                    
                    st.error(f"총 {len(errors_list)}건의 정산 오류 내역이 발견되었습니다.")
                    st.metric(label="🚨 총 초과 청구 금액 합계", value=f"{int(total_excess):,.0f} 원")
                    st.dataframe(styled_errors, use_container_width=True)
                    
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        styled_errors.to_excel(writer, index=False, sheet_name='정산오류내역')
                        worksheet = writer.sheets['정산오류내역']
                        
                        for i, col in enumerate(df_errors.columns):
                            max_len = max(df_errors[col].astype(str).map(len).max(), len(str(col)))
                            adjusted_width = (max_len * 1.8) + 2
                            worksheet.column_dimensions[get_column_letter(i + 1)].width = adjusted_width
                            
                        for row_cells in worksheet.iter_rows(min_row=2):
                            for cell in row_cells:
                                col_name = worksheet.cell(row=1, column=cell.column).value
                                if col_name in ['청구 총금액', '산정 총금액', '초과 청구액']:
                                    try:
                                        if pd.notna(cell.value) and str(cell.value).strip() != '':
                                            cell.value = float(cell.value)
                                            cell.number_format = '#,##0'
                                    except:
                                        pass
                                        
                    excel_data = excel_buffer.getvalue()
                    
                    st.download_button(
                        label="📥 오류 내역 엑셀 다운로드",
                        data=excel_data,
                        file_name='출고배송비_정산오류내역.xlsx',
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                else:
                    st.success("모든 출고 배송비 건이 기준 등급 및 금액 내에서 정상적으로 청구되었습니다.")
                    
                st.subheader("🔍 별도 확인 필요 (SKU 8개 이상)")
                if warnings_list:
                    df_warnings = pd.DataFrame(warnings_list)
                    
                    styled_warnings = df_warnings.style \
                        .set_table_styles([
                            {'selector': 'th', 'props': [('text-align', '
